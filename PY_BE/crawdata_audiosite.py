import openpyxl
import requests
from bs4 import BeautifulSoup
import os
import time
import re

# URL cơ sở
base_url = 'https://audiosite.net/truyen-audio/truyen-do-thi'

# Tên file Excel
excel_file = 'Do_thi_data_detailed.xlsx'

# Số trang tối đa bạn muốn cào
max_pages = 22  # Bạn có thể điều chỉnh số này

# Danh sách chứa thông tin sách từ tất cả các trang
all_book_data = []

# Hàm để trích xuất text từ thẻ span trong thongtin-truyen
def extract_info_from_span(span_text):
    # Nếu có ":" trong văn bản, lấy phần sau dấu ":"
    if ":" in span_text:
        return span_text.split(":", 1)[1].strip()
    return span_text.strip()

# Hàm để lấy thông tin chi tiết từ trang sách
def get_book_details(book_url, book_title):
    try:
        print(f"Đang lấy thông tin chi tiết của sách: {book_title}")
        
        # Gửi yêu cầu HTTP đến trang chi tiết sách
        response = requests.get(book_url)
        response.raise_for_status()
        
        # Phân tích HTML
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Tìm div có class là thongtin-truyen
        info_div = soup.find('div', class_='thongtin-truyen')
        
        if not info_div:
            print(f"Không tìm thấy thông tin chi tiết của sách: {book_title}")
            return {
                "trang_thai": "Không có thông tin",
                "the_loai": "Không có thông tin",
                "tac_gia": "Không có thông tin",
                "giong_doc": "Không có thông tin",
                "so_tap": "Không có thông tin",
                "so_chuong": "Không có thông tin",
                "raw_html": "Không tìm thấy div.thongtin-truyen"
            }
        
        # Lưu lại HTML gốc của div thongtin-truyen
        raw_html = str(info_div)
        
        # Khởi tạo dict để lưu thông tin
        book_info = {
            "trang_thai": "Không có thông tin",
            "the_loai": "Không có thông tin",
            "tac_gia": "Không có thông tin",
            "giong_doc": "Không có thông tin",
            "so_tap": "Không có thông tin",
            "so_chuong": "Không có thông tin",
            "raw_html": raw_html
        }
        
        # Tìm tất cả các thẻ span trong div thongtin-truyen
        spans = info_div.find_all('span')
        
        for span in spans:
            span_text = span.get_text(strip=True)
            
            # Phân tích thông tin từ mỗi span dựa vào nội dung
            if "Trạng thái:" in span_text:
                trang_thai = extract_info_from_span(span_text)
                book_info["trang_thai"] = trang_thai
            
            elif "Thể loại:" in span_text:
                # Lấy tất cả các thẻ 'a' trong span (cho nhiều thể loại)
                the_loai_links = span.find_all('a')
                if the_loai_links:
                    # Lấy tất cả các thể loại và nối lại bằng dấu phẩy
                    the_loai_list = [link.get_text(strip=True) for link in the_loai_links]
                    book_info["the_loai"] = ", ".join(the_loai_list)
                else:
                    book_info["the_loai"] = extract_info_from_span(span_text)
            
            elif "Tác giả:" in span_text:
                # Lấy tất cả các thẻ 'a' trong span (cho nhiều tác giả nếu có)
                tac_gia_links = span.find_all('a')
                if tac_gia_links:
                    # Lấy tất cả các tác giả và nối lại bằng dấu phẩy
                    tac_gia_list = [link.get_text(strip=True) for link in tac_gia_links]
                    book_info["tac_gia"] = ", ".join(tac_gia_list)
                else:
                    book_info["tac_gia"] = extract_info_from_span(span_text)
            
            elif "Giọng đọc:" in span_text:
                # Lấy tất cả các thẻ 'a' trong span (cho nhiều giọng đọc nếu có)
                giong_doc_links = span.find_all('a')
                if giong_doc_links:
                    # Lấy tất cả các giọng đọc và nối lại bằng dấu phẩy
                    giong_doc_list = [link.get_text(strip=True) for link in giong_doc_links]
                    book_info["giong_doc"] = ", ".join(giong_doc_list)
                else:
                    book_info["giong_doc"] = extract_info_from_span(span_text)
            
            elif "Số tập:" in span_text:
                so_tap = re.search(r'Số tập: (\d+)', span_text)
                if so_tap:
                    book_info["so_tap"] = so_tap.group(1)
                else:
                    book_info["so_tap"] = extract_info_from_span(span_text)
            
            elif "Số chương:" in span_text:
                book_info["so_chuong"] = extract_info_from_span(span_text)
        
        return book_info
        
    except Exception as e:
        print(f"Lỗi khi lấy thông tin chi tiết của sách {book_title}: {e}")
        return {
            "trang_thai": "Lỗi",
            "the_loai": "Lỗi",
            "tac_gia": "Lỗi",
            "giong_doc": "Lỗi",
            "so_tap": "Lỗi",
            "so_chuong": "Lỗi",
            "raw_html": f"Lỗi: {str(e)}"
        }

# Lặp qua từng trang
for page_index in range(1, max_pages + 1):
    # Xác định URL của trang hiện tại
    if page_index == 1:
        current_url = base_url
    else:
        current_url = f"{base_url}/page/{page_index}"
    
    print(f"Đang cào dữ liệu từ trang {page_index}: {current_url}")
    
    try:
        # Gửi yêu cầu HTTP để tải trang
        response = requests.get(current_url)
        
        # Kiểm tra nếu có lỗi HTTP
        response.raise_for_status()
        
        # Phân tích nội dung HTML của trang
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Tìm tất cả các book-card trên trang
        book_cards = soup.find_all('article', class_='book-card')
        
        # Nếu không tìm thấy book-card nào, có thể đã hết trang
        if not book_cards:
            print(f"Không tìm thấy sách nào ở trang {page_index}. Dừng lại.")
            break
        
        # Lặp qua từng book-card để lấy thông tin
        for card in book_cards:
            try:
                # Lấy thông tin tiêu đề
                title_element = card.find('h3', class_='book-card-title')
                if title_element and title_element.a:
                    title = title_element.text.strip()
                    
                    # Lấy liên kết (link) đến trang chi tiết
                    link = title_element.a['href']
                    
                    # Lấy hình ảnh (img) của sách
                    img_element = card.find('div', class_='book-front')
                    img = img_element.img['src'] if img_element and img_element.img else "No image"
                    
                    # Lấy thông tin chi tiết từ trang sách (truy cập vào link)
                    book_details = get_book_details(link, title)
                    
                    # Tạo bản ghi đầy đủ với thông tin từ trang chi tiết
                    book_record = [
                        title, 
                        link, 
                        img,
                        book_details["trang_thai"],
                        book_details["the_loai"],
                        book_details["tac_gia"],
                        book_details["giong_doc"],
                        book_details["so_tap"],
                        book_details["so_chuong"],
                        book_details["raw_html"]
                    ]
                    
                    # Lưu thông tin vào danh sách
                    all_book_data.append(book_record)
                    
                    # Tạm dừng giữa các yêu cầu chi tiết để tránh bị chặn
                    time.sleep(1)
                    
            except Exception as e:
                print(f"Lỗi khi xử lý một book-card: {e}")
                continue
        
        print(f"Đã xử lý {len(book_cards)} sách ở trang {page_index}")
        
        # Tạm dừng giữa các trang
        time.sleep(2)
        
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 404:
            print(f"Trang {page_index} không tồn tại (Lỗi 404). Kết thúc việc cào dữ liệu.")
            break
        else:
            print(f"Lỗi HTTP khi tải trang {page_index}: {e}")
            continue
    except Exception as e:
        print(f"Lỗi khi xử lý trang {page_index}: {e}")
        continue

# Thông báo tổng số sách đã thu thập
print(f"Tổng cộng đã thu thập được {len(all_book_data)} sách từ {page_index} trang")

# Kiểm tra xem file Excel đã tồn tại chưa
if os.path.exists(excel_file):
    # Nếu file đã tồn tại, mở file để cập nhật
    print(f"Đang cập nhật file '{excel_file}' hiện có...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Tạo tập hợp các liên kết sách hiện có để tránh trùng lặp
    existing_links = set()
    for row in ws.iter_rows(min_row=2, values_only=True):  # Bắt đầu từ hàng 2 (bỏ qua tiêu đề)
        if row[1]:  # Kiểm tra xem ô link có giá trị không
            existing_links.add(row[1])
    
    # Đếm số sách mới được thêm vào
    new_books_count = 0
    
    # Thêm các sách mới (không trùng lặp) vào file
    for book in all_book_data:
        if book[1] not in existing_links:  # Kiểm tra link chưa có trong file
            ws.append(book)
            existing_links.add(book[1])  # Thêm link vào tập hợp
            new_books_count += 1
    
    print(f"Đã thêm {new_books_count} sách mới vào file.")
    
else:
    # Nếu file chưa tồn tại, tạo file mới
    print(f"File '{excel_file}' chưa tồn tại. Đang tạo file mới...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Books"
    
    # Thêm tiêu đề cột
    headers = [
        "Title", "Link", "Image URL", 
        "Trạng Thái", "Thể Loại", "Tác Giả", 
        "Giọng Đọc", "Số Tập", "Số Chương", 
        "Raw HTML"
    ]
    ws.append(headers)
    
    # Thêm dữ liệu vào sheet
    for book in all_book_data:
        ws.append(book)
    
    print(f"Đã tạo file mới và thêm {len(all_book_data)} sách.")

# Lưu file Excel
wb.save(excel_file)

print(f"Dữ liệu đã được lưu vào file '{excel_file}'.")
